#!/usr/bin/env python3
"""add_media.py — pull demo dashboard screenshots + hardware photos into listings.

Source: ../../iotc-master-catalog.xlsx (sheet "Demos"). Each demo row carries up to
6 "Dashboard N" image URLs (the /IOTCONNECT cloud dashboard) and 5 "Demo Image N"
URLs (real hardware-in-action photos). We match each demo to listing rows in
data/listings.csv and fill two columns:

  Dashboards  pipe-separated dashboard screenshot URLs
  Photos      pipe-separated demo/hardware photo URLs

Matching (each demo -> the listing row(s) it belongs to):
  1. by normalised Name (a demo whose name matches several same-named per-board
     rows attaches to ALL of them, so each board variant shows the demo).
  2. else an explicit ALIAS (demo name -> listing name) for demos named differently
     than their listing.
  3. else by GitHub repo, but ONLY when the repo maps to exactly one listing, or can
     be disambiguated by the demo's Target board(s). Repos shared by several distinct
     demos (e.g. avnet-iotc-mtb-ai-imagimob-rm) are NOT guessed — they need an ALIAS,
     otherwise the demo is left [unmatched] rather than mis-attributed.

Expiring/ephemeral signed image URLs (GitHub private-user-images, ?jwt=,
X-Amz-Signature) are dropped — they 404 once the signature expires.

Re-runnable: rewrites both columns from the workbook each time.
"""
import os, re, csv
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(os.path.dirname(ROOT), "iotc-master-catalog.xlsx")
LISTINGS = os.path.join(ROOT, "data", "listings.csv")
SEP = " | "

def norm(s): return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()
def ghkey(s):
    m = re.search(r"github\.com/([^/]+)/([^/#?]+)", str(s or ""), re.I)
    return (m.group(1).lower() + "/" + m.group(2).lower()) if m else None
def ephemeral(u):
    u = u.lower()
    return ("private-user-images.githubusercontent.com" in u
            or "jwt=" in u or "x-amz-signature" in u)

# demo name -> listing name, for demos whose name matches no listing (and whose
# repo is shared by siblings, so a GitHub guess would be wrong).
ALIAS = {
    "edge ai solutions featuring jetson": "Edge AI Inference Pipeline (Jetson Orin NX)",
    "smart home baby monitor": "PSOC AI-Kit Baby Monitor",
    "psoc6 ai kit human activity": "PSOC AI-Kit Human Activity Detection",
}

# ---- read listings ----
rows = list(csv.reader(open(LISTINGS, newline="", encoding="utf-8")))
H = rows[0]
for col in ("Dashboards", "Photos"):
    if col not in H:
        H.append(col)
idx = {c: H.index(c) for c in H}
data = rows[1:]
for r in data:                                   # pad rows to new width
    while len(r) < len(H): r.append("")

by_name, by_gh = {}, {}                           # both are name/repo -> [rows]
for r in data:
    by_name.setdefault(norm(r[idx["Name"]]), []).append(r)
    gh = ghkey(r[idx["Link"]]) or (("avnet-iotconnect/" + r[idx["Repo"]].lower()) if r[idx["Repo"]] else None)
    if gh: by_gh.setdefault(gh, []).append(r)

for k, v in ALIAS.items():                        # fail loud on a stale alias
    if norm(v) not in by_name:
        raise SystemExit(f"ALIAS target not found: {k!r} -> {v!r}")

def board_set(r):
    return {norm(b) for b in re.split(r"[;,]", r[idx["Boards"]]) if norm(b)}

def resolve(d):
    """Return the list of listing rows this demo belongs to (possibly empty)."""
    nm = norm(d.get("Demo"))
    if nm in by_name: return by_name[nm]
    al = norm(ALIAS.get(nm, ""))
    if al and al in by_name: return by_name[al]
    cands = by_gh.get(ghkey(d.get("Github Link")), [])
    if len(cands) == 1: return cands
    if len(cands) > 1:                            # shared repo: disambiguate by board
        targets = {norm(d.get(f"Target {i}", "")) for i in range(1, 5)}; targets.discard("")
        hits = [r for r in cands if board_set(r) & targets]
        if len(hits) == 1: return hits
        print(f"  [ambiguous-gh] {d.get('Manufacturer')} | {d.get('Demo')} -> "
              f"{len(cands)} listings share the repo; add an ALIAS")
    return []

# ---- read demos ----
wb = load_workbook(XLSX, data_only=True)
ws = wb["Demos"]
hdr = [c.value for c in ws[1]]
demos = [{hdr[i]: (str(rw[i]).strip() if rw[i] is not None else "")
          for i in range(len(hdr)) if hdr[i]}
         for rw in ws.iter_rows(min_row=2, values_only=True) if any(rw)]

# accumulate per listing row (dedup, order-preserving)
dash_by_row, photo_by_row = {}, {}
matched = unmatched = 0
for d in demos:
    dashes = [u for u in (d.get(f"Dashboard {i}", "") for i in range(1, 7))
              if u.startswith("http") and not ephemeral(u)]
    photos = [u for u in (d.get(f"Demo Image {i}", "") for i in range(1, 6))
              if u.startswith("http") and not ephemeral(u)]
    if not (dashes or photos):
        continue
    targets = resolve(d)
    if not targets:
        unmatched += 1
        print(f"  [unmatched] {d.get('Manufacturer')} | {d.get('Demo')} ({len(dashes)+len(photos)} imgs)")
        continue
    matched += 1
    for row in targets:
        rid = id(row)
        for u in dashes:
            dash_by_row.setdefault(rid, [])
            if u not in dash_by_row[rid]: dash_by_row[rid].append(u)
        for u in photos:
            photo_by_row.setdefault(rid, [])
            if u not in photo_by_row[rid]: photo_by_row[rid].append(u)

# ---- write back ----
filled = 0
for r in data:
    rid = id(r)
    r[idx["Dashboards"]] = SEP.join(dash_by_row.get(rid, []))
    r[idx["Photos"]] = SEP.join(photo_by_row.get(rid, []))
    if dash_by_row.get(rid) or photo_by_row.get(rid): filled += 1

csv.writer(open(LISTINGS, "w", newline="", encoding="utf-8"), lineterminator="\n").writerows([H] + data)
print(f"matched {matched} demos -> {filled} listings; {unmatched} unmatched. "
      f"columns: Dashboards, Photos")
