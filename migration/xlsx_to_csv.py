#!/usr/bin/env python3
"""
xlsx_to_csv.py — one-time conversion of the source workbook into git-friendly
CSVs under ../data/ (one file per sheet). After this, the CSVs are the source of
truth and build_site.py reads them; the .xlsx is archived.

CSV conventions: UTF-8 (no BOM), LF line endings, minimal quoting -> clean diffs.
"""
import os, csv
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")
os.makedirs(DATA, exist_ok=True)

wb = load_workbook(os.path.join(ROOT, "iotc-index-catalog.xlsx"), data_only=True)

SHEETS = {"Listings": "listings.csv", "Boards": "boards.csv",
          "Resources": "resources.csv", "Config": "config.csv"}

for sheet, fname in SHEETS.items():
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    # drop trailing empty header columns
    while header and (header[-1] is None or str(header[-1]).strip() == ""):
        header.pop()
    ncol = len(header)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        cells = ["" if (i >= len(r) or r[i] is None) else str(r[i]) for i in range(ncol)]
        if not any(c.strip() for c in cells):
            continue
        rows.append(cells)
    with open(os.path.join(DATA, fname), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)
    print(f"{fname}: {len(rows)} rows x {ncol} cols")
